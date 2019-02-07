import datetime
import collections
import numpy as np
import csv
import io
from Utils.ScientificNotation import convertSI
from wattstrat.simulation.SimulationResultsData import SimulationResultsData
from openpyxl import Workbook as xlWorkbook
from openpyxl.worksheet.dimensions import ColumnDimension

# TODO: remove it when python version = 3.5
from Utils.compatibilities import merge_two_dicts, safe_list_get
from . import SimulationResultsGraph as srg

from . import SimulationResultsDataConfig as SRDConfig

from world import models as wmodels
from world.views import reprojGeom, coordsToBBOX, Proj

from django.contrib.gis.db.models import Union

if __debug__:
    import logging
    logger = logging.getLogger(__name__)
#import pprint
#pp = pprint.PrettyPrinter(indent=4)


class DoNotDisplayData(Exception):
    pass


# Use semicolon instead of comma for the excel dialect
# This is because of how european Excel treat csv files :
# http://stackoverflow.com/questions/7423869/python-csv-writer-wrong-separator
class excel_semicolon(csv.excel):
    delimiter = ';'
csv.register_dialect("excel", excel_semicolon)    

SRDConfig.addAlias("WS.Provider.CSV", "wattstrat.simulation.SimulationResultsProvider.CSVProvider", (), {})
SRDConfig.addAlias("WS.Provider.XLSX", "wattstrat.simulation.SimulationResultsProvider.XLSXProvider", (), {})
SRDConfig.addAlias("WS.Provider.LocalFile", "wattstrat.simulation.SimulationResultsProvider.LocalFileProvider", (), {})
SRDConfig.addAlias("WS.Provider.Zip", "wattstrat.simulation.SimulationResultsProvider.ZipProvider", (), {})
SRDConfig.addAlias("WS.Provider.Concat", "wattstrat.simulation.SimulationResultsProvider.ConcatProvider", (), {})
SRDConfig.addAlias("WS.Provider.ConcatFields", "wattstrat.simulation.SimulationResultsProvider.ConcatFieldsProvider", (), {})

SRDConfig.addAlias("WS.Provider.Graphic.map", "wattstrat.simulation.SimulationResultsProvider.MapChartProvider", (), {"graphType": "map"})
SRDConfig.addAlias("WS.Provider.Graphic.lineChart", "wattstrat.simulation.SimulationResultsProvider.TimeSerialChartProvider", (), {"graphType": "lineChart"})
SRDConfig.addAlias("WS.Provider.Graphic.stackedLineChart", "wattstrat.simulation.SimulationResultsProvider.TimeSerialChartProvider", (), {"graphType": "stackedLineChart"})
SRDConfig.addAlias("WS.Provider.Graphic.pieChart", "wattstrat.simulation.SimulationResultsProvider.PieChartProvider", (), {"graphType": "pieChart"})
SRDConfig.addAlias("WS.Provider.Graphic.histofreqChart", "wattstrat.simulation.SimulationResultsProvider.HistoChartProvider", (), {"graphType": "histofreqChart"})
SRDConfig.addAlias("WS.Provider.Graphic.histodateChart", "wattstrat.simulation.SimulationResultsProvider.TimeSerialChartProvider", (), {"graphType": "histodateChart"})


class ContainerProvider(SimulationResultsData):
    def __init__(self, *args, inputs=None, fileKey="filename", dataKey="data", uuid=None, configuration=None, **kwargs):
        super().__init__(*args, inputs=inputs, uuid=uuid, configuration=configuration, **kwargs)
        
        self._Container = None
        self._fileIndex = 0
        self._dataKey = dataKey
        self._fileKey = fileKey
        self._alreadyIter = False

    def __iter__(self):
        self._alreadyIter = False
        return self

    def __next__(self):
        if self._alreadyIter:
            # Iter just once
            raise StopIteration()
        self._alreadyIter = True
        while True:
            try:
                datas = super().__next__()
                if len(datas) == 0:
                    break
                for data in datas:
                    if self._fileKey in data:
                        if self._dataKey in data:
                            self._Container.write(data[self._dataKey], filename=data[self._fileKey])
                    else:
                        filename = self._inferFilename(data)
                        self._Container.write(json.dumps(data), filename=filename)
            except StopIteration:
                break
        
        return [self._Container.data()]
    
    def _inferFilename(self, data):
        self._fileIndex += 1
        return "file-%d.json" % (self._fileIndex)
    
class ZipProvider(ContainerProvider):
    def __init__(*args, zipname="file.zip",password=None, fileKey="filename", dataKey="data", **kwargs):
        super().__init__(*args, fileKey=fileKey, dataKey=dataKey, **kwargs)
        self._Container = ZipContainer(zipname=zipname, fileKey=fileKey, dataKey=dataKey, password=password)

class ConcatProvider(ContainerProvider):
    def __init__(*args, fileKey="filename", dataKey="data", **kwargs):
        super().__init__(*args, fileKey=fileKey, dataKey=dataKey, **kwargs)
        self._Container = ConcatContainer()
        
class ConcatFieldsProvider(ContainerProvider):
    def __init__(*args, fields=['data'], fileKey="filename", dataKey="data", **kwargs):
        super().__init__(*args, fileKey=fileKey, dataKey=dataKey, **kwargs)
        self._Container = ConcatFieldsContainer(fields=fields)

class LocalFileProvider(ContainerProvider):
    def __init__(*args, fileKey="filename", dataKey="data", **kwargs):
        super().__init__(*args, fileKey=fileKey, dataKey=dataKey, **kwargs)
        self._Container = LocalFileContainer(fileKey=fileKey, dataKey=dataKey)

class CSVProvider(SimulationResultsData):
    def __init__(self, *args, inputs=None, format="excel", filename="results",uuid=None, configuration=None, **kwargs):
        super().__init__(*args, inputs=inputs, uuid=uuid, configuration=configuration, **kwargs)
        self._format = "excel"
        self._dialect = "excel" if self._format == "excel" else "unix"
        self._filename = filename

    def __iter__(self): 
        return self

    def __next__(self):
        iobuf = io.StringIO()
        csvWriter = csv.writer(iobuf, dialect=self._dialect)
        datas = super().__next__()
        if len(datas) == 0:
            raise StopIteration()
        
        header  = self.build_header(datas)
        csvWriter.writerow(header)

        if datas[0]['precision'] in ['y', 'm', 'w', 'd']:
            filename = '%s.csv' % (self._filename)
        else:
            filename = '%s-%d.csv' % (self._filename, datas[0]["year"])


        lbl = {"label":'Nom de la variable', "unit": "Unité", "geoLabel":'Nom géographique', "geocode": "geocode"}
        for ind in ["label", "unit", "geoLabel", "geocode"]:
            row = [lbl[ind]]
            for data in datas:
                cData = data.get(ind, "")
                row.append(cData)
            csvWriter.writerow(row)

        # csvWriter.writerow(["date"])

        try:
            for indV in range(len(datas[0]["values"])):
                row = [self._formatTime(indV, datas[0])]
                row.extend(map(lambda x: safe_list_get(x["values"], indV, 0), datas))
                csvWriter.writerow(row)
        except StopIteration:
            pass
        
        return {
            'filename': filename,
            'data': iobuf.getvalue(),
        }

    def build_header(self, datas):
        return []

    def _formatTime(self, indV, out):
        date_formats = {
            'y': "%Y",
            'm': "%b %Y",
            'w': "%Y-Week %W",
            'd': "%Y-%m-%d",
            '6h': "%Y-%m-%d %Hh",
            'h': "%Y-%m-%d %Hh",
        }

        # From def _defaultTimeFormat(indV, out, precision):

        # Change time format to format for CSV
        precision = out.get("precision", "h")
        end = out.get('end')
        formattedTime = date_formats.get(precision)
        if formattedTime is None:
            formattedTime = date_formats['h']
        date = out.get("begin")
        if date is None:
            date = datetime.datetime(out['year'], 1, 1)

        step = out.get("step")
        if step is None:
            if precision == 'h':
                step = datetime.timedelta(hours=indV)
            elif precision == 'd':
                step = datetime.timedelta(days=indV)
            elif precision == 'w':
                step = datetime.timedelta(weeks=indV)
            else:
                year = date.year
                month = date.month
                if precision == 'm':
                    year += int(indV / 12)
                    month += (indV % 12)
                else:
                    year += indV
                current = datetime.datetime(year, month, 1, 0, 0, 0)
        else:
            # Get step from variable, shoul dbe a function
            step, current = step(indV)
            
        if step is not None:
            current = date + step

        if end is not None:
            if current >= end:
                raise StopIteration()

        return current.strftime(formattedTime)

class XLSXProvider(SimulationResultsData):
    MAX_COLUMN_SIZELETTER = 150
    def __init__(self, *args, inputs=None, allInOne=False, filename="results",uuid=None, configuration=None, **kwargs):
        super().__init__(*args, inputs=inputs, uuid=uuid, configuration=configuration, **kwargs)
        self._filename = filename
        self._all = allInOne
        if allInOne:
            self._iobuf = io.BytesIO()
            self._xlsxWriter = xlWorkbook()
            self._first=True
            self._saved = False

    def _saveBench(self, w, iobuf, filename):
        w.save(iobuf)
        return {
            'filename': filename,
            'data': iobuf.getvalue(),
        }

    def __iter__(self): 
        return self

    def __next__(self):
        while True:
            if self._all:
                if self._first:
                    self._first=False
                    ws = self._xlsxWriter.active
                else:
                    ws = self._xlsxWriter.create_sheet()
                iobuf=self._iobuf
                xlsxWriter = self._xlsxWriter
                
            else:    
                iobuf = io.BytesIO()
                xlsxWriter = xlWorkbook()
                ws=xlsxWriter.active
    
            try:
                datas = super().__next__()
                if len(datas) == 0:
                    if self._all and not self._saved:
                        self._saved = True
                        return self._saveBench(xlsxWriter, iobuf, '%s.xlsx' % (self._filename))
                    raise StopIteration()
            except StopIteration as e:
                if self._all and not self._saved:
                    self._saved = True
                    return self._saveBench(xlsxWriter, iobuf, '%s.xlsx' % (self._filename))
                raise e
    
            if datas[0]['precision'] in ['y', 'm', 'w', 'd']:
                ws.title = "Data"
            else:
                ws.title = '%d' % (datas[0]["year"])
    
            if datas[0]['precision'] in ['y', 'm', 'w', 'd']:
                filename = '%s.xlsx' % (self._filename)
            else:
                filename = '%s-%d.xlsx' % (self._filename, datas[0]["year"])
            
            header  = self.build_header(datas)
            if header is not None and len(header)>0:
                ws.append(header)
    
            lbl = {"label":'Nom de la variable', "unit": "Unité", "geoLabel":'Nom géographique', "geocode": "geocode"}

            for ind in ["label", "unit", "geoLabel", "geocode"]:
                row = [lbl[ind]]
                for data in datas:
                    cData = data.get(ind, "")
                    row.append(cData)
                ws.append(row)
    
            try:
                for indV in range(len(datas[0]["values"])):
                    row = [self._formatTime(indV, datas[0])]
                    row.extend(map(lambda x: safe_list_get(x["values"], indV, 0), datas))
                    ws.append(row)
            except StopIteration:
                pass
            self._resize_cells(ws)
            if not self._all:
                break
        return self._saveBench(xlsxWriter, iobuf, filename)

    def _resize_cells(self, ws):
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), min(len(str(cell.value)), self.MAX_COLUMN_SIZELETTER)))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value#*1.1

    def build_header(self, datas):
        return []

    def _formatTime(self, indV, out):
        date_formats = {
            'y': "%Y",
            'm': "%b %Y",
            'w': "%Y-Week %W",
            'd': "%Y-%m-%d",
            '6h': "%Y-%m-%d %Hh",
            'h': "%Y-%m-%d %Hh",
        }

        # From def _defaultTimeFormat(indV, out, precision):

        # Change time format to format for CSV
        precision = out.get("precision", "h")
        end = out.get('end')
        formattedTime = date_formats.get(precision)
        if formattedTime is None:
            formattedTime = date_formats['h']
        date = out.get("begin")
        if date is None:
            date = datetime.datetime(out['year'], 1, 1)

        step = out.get("step")
        if step is None:
            if precision == 'h':
                step = datetime.timedelta(hours=indV)
            elif precision == 'd':
                step = datetime.timedelta(days=indV)
            elif precision == 'w':
                step = datetime.timedelta(weeks=indV)
            else:
                year = date.year
                month = date.month
                if precision == 'm':
                    year += int(indV / 12)
                    month += (indV % 12)
                else:
                    year += indV
                current = datetime.datetime(year, month, 1, 0, 0, 0)
        else:
            # Get step from variable, shoul dbe a function
            step, current = step(indV)
            
        if step is not None:
            current = date + step

        if end is not None:
            if current >= end:
                raise StopIteration()

        return current.strftime(formattedTime)

# TODO: split class for styling AMcharts
# orderData: take a tuple and give the key to sort the dict


class SerialChartProvider(SimulationResultsData):

    def __init__(self, *args, inputs=None, funCategorie=None, addKeyDictAlone=False, orderData=None,
                 title=None, theme=None,
                 abscisse=None,
                 style=None,
                 axis=False, graphType = None,
                 uuid=None, configuration=None, **kwargs):
        self._orderData = orderData
        self._funCategorie = TimeSerialChartProvider._defaultCat if funCategorie is None else funCategorie
        self._addKeyDictAlone = addKeyDictAlone

        self._title = title
        self._theme = "light" if theme is None else theme

        self._axis = axis
        self._graphType = graphType
        if abscisse is None:
            self._abscisse = {
                'type': 'other',
                'field': 'key',
            }
        else:
            self._abscisse = abscisse

        self._style = style
        # Keep outputs
        super().__init__(*args, inputs=inputs, nOutputs=None, uuid=uuid, configuration=configuration, **kwargs)

    @staticmethod
    def _defaultCat(ind, keyDict, out, **kwargs):
        cat = []
        if keyDict is not None and kwargs.get('key_dict_alone', False):
            return keyDict

        for k in ["field", "shortLabel", "label", "varname"]:
            cat.append(out.get(k))

        cat.append("cat-%d" % (ind))

        categorie = None
        for c in cat:
            if c is not None:
                categorie = c
                break

        if keyDict is None:
            return categorie
        else:
            return "%s-%s" % (categorie, keyDict)

    def __iter__(self):
        return self

    def __next__(self):
        dRet = {}
        values = []
        conf = []
        outs = super().__next__()
        for ind in range(len(outs)):
            self._updateData(ind, outs[ind], dRet)
            conf.append(outs[ind].copy())
            del conf[ind]["values"]
        if self._orderData is None:
            values = [merge_two_dicts({'key': k},v) for k, v in dRet.items()]
        else:
            values = [merge_two_dicts({'key': k},v) for k, v in sorted(dRet.items(), key=self._orderData)]

        conf = self._updateConf(conf)
        conf["values"] = values
        if self._axis:
            axis = srg.Chart.createAxis(conf)
            conf["axis"] = axis
        conf["graphType"] = self._graphType
        self._outs = [conf]
        self._actualNumberOutput = 1

        return self._outs

    def _updateConf(self, conf):
        config = {}
        # Serial with theme
        config["type"] = "serial"
        config["theme"] = self._theme
        # Define all graphs

        config["y"] = [dict(self._modifStyle(param, ind)) for ind, param in enumerate(conf)]
        # Define abscisse
        config["x"] = self._abscisse

        # Add titles
        if self._title is None:
            config["titles"] = []
        else:
            config["titles"] = [
                {
                    'text': self._title,
                    'index': 0,
                    'style': {
                        'bold': False,
                        'color': '#000000',
                        'alpha': 1.0,
                        'fontSize': 15,
                    }
                }
            ]
        return config

    def _updateData(self, ind, out, ret):
        val = out['values']
        if type(val) is dict:
            for k, v in val.items():
                self._updateValuesData(ind, k, v, out, ret)
        else:
            self._updateValuesData(ind, None, val, out, ret)

    def _updateValuesData(self, ind, keyDict, val, out, ret):
        # List of 1 value = 1 value
        if type(val) is list and len(val) == 1:
            val = val[0]
        # Static cat for one variable
        cat = self._funCategorie(ind, keyDict, out, key_dict_alone=self._addKeyDictAlone)
        d = ret.get(cat, {})
        d[cat] = {'key': cat, 'values': val}
        ret[cat] = d

    def _modifStyle(self, param, index):
        if self._style is not None and len(self._style) > index and self._style[index] is not None:
            param["style"] = self._style[index]
        return param

# formatTime: Given a precision, format time label
# funCategorie: function to extract categorie name from variable. Default
# : values = dict: key of dict, field, shortLabel, label, varname,
# "cat-<position>"


class TimeSerialChartProvider(SerialChartProvider):
    PRECISION_FROM_LEN = [
        (8784, 'h'),
        (365, 'd'),
        (52, 'w'),
        (12, 'm'),

    ]
    FORMAT_FROM_PRECISON = {
        'h': '%Y%m%d%H',
        'd': '%Y%m%d',
        'w': '%Y%m%d',
        'm': '%Y%m',
        'y': '%Y'
    }
    FORMATAMCHARTS_FROM_PRECISON = {
        'h': {
            'balloon': 'DD/MM/YYYY JJh',
            'format': 'YYYYMMDDJJ',
            'minPeriod': 'hh',
        },
        'd': {
            'balloon': 'DD/MM/YYYY',
            'format': 'YYYYMMDD',
            'minPeriod': 'DD',
        },
        'w': {
            'balloon': '[W] YYYY',
            'format': 'YYYYMMDD',
            'minPeriod': '7DD',
        },
        'm': {
            'balloon': 'MMM YYYY',
            'format': 'YYYYMM',
            'minPeriod': 'MM',
        },
        'y': {
            'balloon': 'YYYY',
            'format': 'YYYY',
            'minPeriod': 'YYYY',
        },
    }

    def __init__(self, *args, inputs=None, formatTime=None, funCategorie=None, addKeyDictAlone=False, orderData=None,
                 title=None, theme=None,
                 style=None,
                 uuid=None, configuration=None, **kwargs):
        self._formatTime = TimeSerialChartProvider._defaultTimeFormat if formatTime is None else formatTime
        # Order key by date
        super().__init__(*args, inputs=inputs,
                         funCategorie=funCategorie, addKeyDictAlone=addKeyDictAlone, orderData=lambda z: z[0],
                         title=title, theme=theme,
                         abscisse=None,
                         style=style,
                         uuid=uuid, configuration=configuration, **kwargs)

    def _updateConf(self, conf):
        # super could modify conf via _modifStyle
        precision = conf[0].get('precision', 'h')
        AM = conf[0].get('AMCHARTS')

        config = super()._updateConf(conf)
        if not all([param["precision"] == conf[0]['precision'] for param in conf]):
            if __debug__:
                logger.error("Not all at same precision")

        if AM is None:
            config["x"] = {
                'type': 'date',
                'field': 'key',
                'format': TimeSerialChartProvider.FORMATAMCHARTS_FROM_PRECISON[precision]['format'],
                'minPeriod': TimeSerialChartProvider.FORMATAMCHARTS_FROM_PRECISON[precision]['minPeriod'],
                'formatBalloon': TimeSerialChartProvider.FORMATAMCHARTS_FROM_PRECISON[precision]['balloon'],
            }
        else:
            config["x"] = {
                'type': 'date',
                'field': 'key',
                'format': AM['format'],
                'minPeriod': AM['minPeriod'],
                'formatBalloon': AM['balloon'],
            }
        return config

    @staticmethod
    def _defaultTimeFormat(indV, out, precision):
        formattedTime = out.get("timeformat")
        end = out.get('end')
        if formattedTime is None:
            formattedTime = TimeSerialChartProvider.FORMAT_FROM_PRECISON[precision]
        date = out.get("begin")
        if date is None:
            date = datetime.datetime(out['year'], 1, 1)
        step = out.get("step")
        if step is None:
            if precision == 'h':
                step = datetime.timedelta(hours=indV)
            elif precision == 'd':
                step = datetime.timedelta(days=indV)
            elif precision == 'w':
                step = datetime.timedelta(weeks=indV)
            else:
                year = date.year
                month = date.month
                if precision == 'm':
                    year += int(indV / 12)
                    month += (indV % 12)
                else:
                    year += indV
                current = datetime.datetime(year, month, 1, 0, 0, 0)
        else:
            # Get step from variable, shoul dbe a function
            step, current = step(indV)

        if step is not None:
            current = date + step
        if end is not None:
            if current >= end:
                raise StopIteration()
        return current.strftime(formattedTime)

    @staticmethod
    def _getPrecision(out, val):
        prec = out.get("precision")

        lprec = len(val)
        if prec is None:
            for p in TimeSerialChartProvider.PRECISION_FROM_LEN:
                if (lprec % p[0]) == 0:
                    prec = p[1]
                    break
            if prec is None:
                prec = 'y'
        return prec

    def _updateValuesData(self, ind, keyDict, val, out, ret):
        if type(val) is not list:
            val = [val]

        precision = TimeSerialChartProvider._getPrecision(out, val)
        # Static cat for one variable
        cat = self._funCategorie(ind, keyDict, out, precision=precision)
        try:
            for indV in range(len(val)):
                # Key is time
                try:
                    lbl = self._formatTime(indV, out, precision)
                except DoNotDisplayData:
                    pass
                d = ret.get(lbl, {})
                d[cat] = val[indV]
                ret[lbl] = d
        except StopIteration:
            pass        

    def _modifStyle(self, param, index):
        param = super()._modifStyle(param, index)
        # Remove specifics keys
        for k in ["AMCHARTS", "step", "timeformat"]:
            try:
                del param[k]
            except KeyError:
                pass
        return param


# One input per variable,  values : dict{geocode => value}


class MapChartProvider(SimulationResultsData):
    Proj=4326
    def __init__(self, *args, inputs=None, extractLabel=None, extractData=None, axis=False, graphType=None, uuid=None, configuration=None, **kwargs):
        self._extractLabel = MapChartProvider._defaultLabel if extractLabel is None else extractLabel
        self._extractData = MapChartProvider._defaultData if extractData is None else extractData
        self._axis = axis
        self._graphType = graphType
        super().__init__(*args, inputs=inputs, nOutputs=None, uuid=uuid, configuration=configuration, **kwargs)

    def __iter__(self):
        return self

    @staticmethod
    def _defaultLabel(ind, out):
        lbl = []
        for k in ["field", "shortLabel", "label", "varname"]:
            lbl.append(out.get(k))

        lbl.append("var-%d" % (ind))

        label = None
        for l in lbl:
            if l is not None:
                label = l
                break

        return label

    @staticmethod
    def _defaultData(key, val):
        # By default, sum data
        return np.sum(val)

    def __next__(self):
        ret = {}
        outs = super().__next__()
        retFields = []
        retLabel = []
        for ind in range(len(outs)):
            lbl = self._extractLabel(ind, outs[ind])
            retFields.append(lbl)
            retLabel.append(outs[ind]["label"])
            if type(outs[ind]["values"]) is not dict:
                raise TypeError("In carto provider, input should be a dict of value : %s", type(outs[ind]["values"]))
            dValues = {k: self._extractData(k, v) for k, v in outs[ind]['values'].items()}
            values = list(dValues.values())
            try:
                (minV, maxV) = (np.min(values), np.max(values))
            except ValueError:
                minV = 0
                maxV = 0

            ret[lbl] = {'values': dValues, 'min': minV, 'max': maxV, 'unit': outs[ind]['unit']}
            if 'batiments' in outs[ind]:
                bValues = {k: self._extractData(k, v) for k, v in outs[ind]['batiments'].items()}
                batVals = list(bValues.values())
                try:
                    (bminV, bmaxV) = (np.min(batVals), np.max(batVals))
                except ValueError:
                    bminV = 0
                    bmaxV = 0

                ret[lbl].update({'batiments': bValues, 'batMin': bminV, 'batMax': bmaxV})

            # Construct aggreation of all geocode
            bboxCom = wmodels.OSMCommunes.objects.filter(geocode__in=list(outs[ind]["values"])).aggregate(area=Union('geom'))['area']
            bboxDept = wmodels.OSMDepartements.objects.filter(geocode__in=list(outs[ind]["values"])).aggregate(area=Union('geom'))['area']
            if bboxCom is None:
                bbox=bboxDept
            elif bboxDept is None:
                bbox=bboxCom
            else:
                bbox=bboxCom
                bbox.union(bboxDept)
            if bbox is not None:
                bbox = bbox.envelope
                bbox.transform(reprojGeom[self.Proj])
                ret[lbl]["bbox"] = bbox.wkt

        if self._axis:
            axis = srg.Chart.createAxis(ret)
            ret["axis"] = axis
        
        ret["graphType"] = self._graphType
        ret["parametersList"] = retFields
        ret["parametersLabel"] = retLabel
        self._outs = [ret]
        self._actualNumberOutput = 1
        return self._outs

# Histogram : 1 input only
# TODO: Use BarChar styling


class HistoChartProvider(SimulationResultsData):

    def __init__(self, *args, inputs=None,
                 title = None, theme = None,
                 style = None,
                 percent = True,
                 bins = "doane",
                 axis=False, graphType=None,
                 uuid=None, configuration=None, **kwargs):

        self._title = title
        self._theme = "light" if theme is None else theme
        self._style = style
        self._percent = percent
        self._bins = bins
        self._axis = axis
        self._graphType = graphType
        super().__init__(*args, inputs=inputs, nOutputs=None, uuid=uuid, configuration=configuration, **kwargs)

    def __iter__(self):
        return self

    def __next__(self):
        ret = []
        outs = super().__next__()

        if len(outs) == 0:
            raise StopIteration()

        for out in outs:
            if type(out["values"]) is not list:
                raise TypeError("histoChart should have list of values, not %s" % (type(out["values"])))

            numValue = len(out["values"])
            histo, edges = np.histogram(out["values"], bins=self._bins)
            values = [ { 'value': (val*100)/numValue if self._percent else val,
                         'edge_b': edgb, 'edge_e': edge,
                         'label': "[%s%s,%s%s[" % ( convertSI(edgb, 2), out.get("unit", "Wh"), convertSI(edge, 2), out.get("unit", "Wh"))
            } for val, edgb, edge in zip(histo, edges[0:-1], edges[1:])]

            conf = self._updateConf(out)
            conf["values"] = values
            if self._axis:
                axis = srg.Chart.createAxis(conf)
                conf["axis"] = axis
            conf["graphType"] = self._graphType
            ret.append(conf)
        self._actualNumberOutput = len(ret)
        return ret

    # From SerialProvider
    def _updateConf(self, conf):
        config = {}
        # Serial with theme
        config["type"] = "serial"
        config["theme"] = self._theme
        # Define all graphs

        config["y"] = [merge_two_dicts({'unit': '%' if self._percent else "nb", 'field': 'value', 'varname': conf["varname"],
                                        'label': conf['label']},self._modifStyle(conf))]
        # Define abscisse
        config["x"] = {
            'type': 'other',
            'field': 'label',
        }

        # Add titles
        if self._title is None:
            config["titles"] = []
        else:
            config["titles"] = [
                {
                    'text': self._title,
                    'index': 0,
                    'style': {
                        'bold': False,
                        'color': '#000000',
                        'alpha': 1.0,
                        'fontSize': 15,
                    }
                }
            ]
        return config

    def _modifStyle(self, param):
        if self._style is not None:
            param["style"] = self._style
        for clean in ["field", "varname", "label", "unit"]:
            try:
                del param[clean]
            except KeyError:
                pass

        return param


class PieChartProvider(SimulationResultsData):

    def __init__(self, *args, inputs=None, funCategorie=None, addKeyDictAlone=False, orderData=None,
                 title=None, theme=None,
                 axis=False, graphType=None,
                 uuid=None, configuration=None, **kwargs):

        self._title = title
        self._theme = "light" if theme is None else theme

        self._axis = axis
        self._graphType = graphType
        # Keep outputs
        super().__init__(*args, inputs=inputs, nOutputs=None, uuid=uuid, configuration=configuration, **kwargs)

        # Iterator
        self._indexValues = None

    def __iter__(self):
        return self

    def __next__(self):
        outs = None
        if self._indexValues is None:
            outs = super().__next__()
            self._indexValues = 0
        else:
            outs = self._outs
            self._indexValues += 1

        ret = []
        for out in outs:
            if type(out["values"]) in [int, float] and self._indexValues == 0:
                ret.append(self._formatOut(out), out["values"])
            elif type(out["values"]) is list:
                try:
                    ret.append(self._formatData(out, self._indexValues))
                except IndexError:
                    # no more values
                    pass
            else:
                if __debug__:
                    logger.debug("Could only treat list/fload/int in values for PieChart")
                pass
        if len(ret) == 0:
            self._indexValues = None
            return self.__next__()

        config = self._updateConf(outs)
        config["values"] = ret
        if self._axis:
            axis = srg.Chart.createAxis(config)
            config["axis"] = axis
        config["graphType"] = self._graphType
        self._actualNumberOutput = 1
        return [config]

    def _formatData(self, out, index):
        value = out["values"][index]
        return {'key': out["label"],
                'value': value,
                'unit': out["unit"],
                }

    def _updateConf(self, outs):
        config = {}
        # Serial with theme
        config["type"] = "pie"
        config["theme"] = self._theme
        config["valueField"] = "value"
        config["titleField"] = "key"
        # Add titles
        if self._title is None:
            config["titles"] = []
        else:
            config["titles"] = [
                {
                    'text': self._title,
                    'index' : 0,
                    'style': {
                        'bold': False,
                        'color': '#000000',
                        'alpha': 1.0,
                        'fontSize': 15,
                    }
                }
            ]
        return config
